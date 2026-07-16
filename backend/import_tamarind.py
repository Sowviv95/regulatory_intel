"""
Tamarind Batch 1 importer.

Imports structured extraction outputs from the Tamarind PoC into the
Regulatory Intelligence SQLite database.

Usage:
    python -m backend.import_tamarind --input "path/to/Batch1.xlsx"
    python -m backend.import_tamarind --input "path/to/Batch1.xlsx" --dry-run
    python -m backend.import_tamarind --input "path/to/Batch1.xlsx" --reset-imported

Run from the project root (D:\\Regulatory Intelligence).
"""
import argparse
import ast
import json
import os
import sys
from datetime import datetime
from pathlib import Path

# Ensure backend is importable when run as `python -m backend.import_tamarind`
sys.path.insert(0, str(Path(__file__).parent))

from database import get_connection, init_db

# ---------------------------------------------------------------------------
# Country metadata lookup
# ---------------------------------------------------------------------------

COUNTRY_META = {
    "Taiwan": {"flag": "\U0001f1f9\U0001f1fc", "language": "ZH-TW"},
    "South Korea": {"flag": "\U0001f1f0\U0001f1f7", "language": "KO"},
    "Denmark": {"flag": "\U0001f1e9\U0001f1f0", "language": "DA"},
    "Poland": {"flag": "\U0001f1f5\U0001f1f1", "language": "PL"},
    "Finland": {"flag": "\U0001f1eb\U0001f1ee", "language": "FI"},
    "Vietnam": {"flag": "\U0001f1fb\U0001f1f3", "language": "VI"},
    "United Arab Emirates": {"flag": "\U0001f1e6\U0001f1ea", "language": "AR"},
    "Albania": {"flag": "\U0001f1e6\U0001f1f1", "language": "SQ"},
    "Moldova": {"flag": "\U0001f1f2\U0001f1e9", "language": "RO"},
    "United Kingdom": {"flag": "\U0001f1ec\U0001f1e7", "language": "EN"},
    "United States": {"flag": "\U0001f1fa\U0001f1f8", "language": "EN"},
    "European Union": {"flag": "\U0001f1ea\U0001f1fa", "language": "EN"},
}

# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------


def parse_products(raw) -> str:
    """Parse tobacco products from mixed formats to semicolon-separated string."""
    if not raw:
        return ""
    s = str(raw).strip()
    # Format B: already semicolon-separated plain text
    if not s.startswith("["):
        return s
    # Format A: list of dicts with picklist_term or label keys
    try:
        items = ast.literal_eval(s)
        labels = []
        for item in items:
            if isinstance(item, dict):
                if "picklist_term" in item:
                    pt = item["picklist_term"]
                    if isinstance(pt, dict):
                        labels.extend(pt.keys())
                    else:
                        labels.append(str(pt))
                elif "label" in item:
                    labels.append(item["label"])
                else:
                    labels.extend(item.keys())
        return "; ".join(labels) if labels else s
    except Exception:
        return s


def parse_products_confidence(raw) -> float:
    """Extract average confidence from product relevance scores."""
    if not raw:
        return 90.0
    s = str(raw).strip()
    if not s.startswith("["):
        return 90.0
    try:
        items = ast.literal_eval(s)
        scores = []
        for item in items:
            if isinstance(item, dict):
                if "picklist_term" in item:
                    pt = item["picklist_term"]
                    if isinstance(pt, dict):
                        scores.extend(pt.values())
                elif "relevance" in item:
                    scores.append(item["relevance"])
                else:
                    scores.extend(item.values())
        if scores:
            return round(sum(scores) / len(scores) * 100, 1)
    except Exception:
        pass
    return 90.0


def parse_source_type(raw) -> str:
    """Extract source type label from mixed formats."""
    if not raw:
        return "Unknown"
    s = str(raw).strip()
    if not s.startswith("["):
        return s
    try:
        items = ast.literal_eval(s)
        if items and isinstance(items[0], dict):
            return list(items[0].keys())[0]
    except Exception:
        pass
    return s


def parse_source_type_confidence(raw) -> float:
    """Extract confidence from source type."""
    if not raw:
        return 90.0
    s = str(raw).strip()
    if not s.startswith("["):
        return 90.0
    try:
        items = ast.literal_eval(s)
        if items and isinstance(items[0], dict):
            return round(list(items[0].values())[0] * 100, 1)
    except Exception:
        pass
    return 90.0


def clean_text(val) -> str:
    """Strip stray leading quotes and whitespace."""
    if val is None:
        return ""
    s = str(val).strip()
    if s.startswith("'") and not s.startswith("''"):
        s = s[1:]
    return s


def derive_topic(products_str: str) -> str:
    """Derive a topic label from products impacted."""
    p = products_str.lower()
    if "e-cigarette" in p or "vaping" in p:
        return "ENDS Regulation"
    if "heated tobacco" in p:
        return "HTP Regulation"
    if "nicotine pouch" in p:
        return "Nicotine Pouches"
    if "cigarette" in p:
        return "Tobacco Control"
    return "Tobacco & Nicotine"


# ---------------------------------------------------------------------------
# Excel reader
# ---------------------------------------------------------------------------


def read_batch1_xlsx(path: str) -> tuple[list[dict], list[dict]]:
    """Read Batch 1 Excel file. Returns (records, errors)."""
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl is required. Install with: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, read_only=True)
    # Find the data sheet
    sheet_name = None
    for name in wb.sheetnames:
        if "batch" in name.lower() and "data" in name.lower():
            sheet_name = name
            break
    if not sheet_name:
        # Fall back to first sheet
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    records = []
    errors = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        vals = {headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)}
        # Skip empty rows
        if not vals.get("Jurisdiction name") and not vals.get("jurisdiction_name"):
            continue
        # Normalize column names
        record = normalize_columns(vals)
        # Validate required fields
        missing = [f for f in ["jurisdiction", "source_name", "title", "url"] if not record.get(f)]
        if missing:
            errors.append({
                "row": row_idx,
                "error": f"Missing required fields: {', '.join(missing)}",
                "data": {k: str(v)[:100] for k, v in record.items()},
            })
            continue
        record["_row"] = row_idx
        records.append(record)

    wb.close()
    return records, errors


# Column name normalization (handles both Batch 1 and Albania/Moldova schemas)
COLUMN_MAP = {
    "Jurisdiction name": "jurisdiction",
    "jurisdiction_name": "jurisdiction",
    "Official name of the source": "source_name",
    "official_source_name": "source_name",
    "official_name": "source_name",
    "Descriptive name": "title",
    "descriptive_name": "title",
    "Summary of the source": "summary",
    "summary": "summary",
    "Summary": "summary",
    "Tobacco Products impacted": "products",
    "tobacco_products_impacted": "products",
    "impacted_prouct": "products",
    "Sub-Product impacted": "sub_products",
    "sub_products_impacted": "sub_products",
    "impacted_Sub_Product": "sub_products",
    "Type of Source": "source_type",
    "type_of_source": "source_type",
    "Status": "status",
    "status": "status",
    "Impact on sector": "impact",
    "impact_on_sector": "impact",
    "Comment: impact on sector": "impact_comment",
    "impact_comment": "impact_comment",
    "Who proposed it?": "proposer",
    "proposed_by": "proposer",
    "who_proposed_it": "proposer",
    "Likelihood of passing": "likelihood",
    "likelihood_of_passing": "likelihood",
    "Comment: likelihood of passing": "likelihood_comment",
    "likelihood_comment": "likelihood_comment",
    "Relevant dates and descripition": "dates",
    "relevant_dates": "dates",
    "relevant_dates_and_descripition": "dates",
    "URL": "url",
    "url": "url",
}


def normalize_columns(vals: dict) -> dict:
    """Normalize column names from any Tamarind schema variant."""
    result = {}
    for key, value in vals.items():
        if key in COLUMN_MAP:
            result[COLUMN_MAP[key]] = value
    return result


# ---------------------------------------------------------------------------
# Import logic
# ---------------------------------------------------------------------------


def build_fields(record: dict) -> list[dict]:
    """Build the 12 regulation fields from a Tamarind record."""
    products_str = parse_products(record.get("products"))
    products_conf = parse_products_confidence(record.get("products"))
    sub_products_str = parse_products(record.get("sub_products"))
    source_type_str = parse_source_type(record.get("source_type"))
    source_type_conf = parse_source_type_confidence(record.get("source_type"))

    return [
        {"category": "Metadata", "field_name": "Jurisdiction",
         "value": record["jurisdiction"], "confidence": 99,
         "evidence": record["jurisdiction"]},
        {"category": "Metadata", "field_name": "Source Name",
         "value": record["source_name"], "confidence": 97,
         "evidence": record["source_name"]},
        {"category": "Metadata", "field_name": "Source Type",
         "value": source_type_str, "confidence": source_type_conf,
         "evidence": str(record.get("source_type", ""))},
        {"category": "Metadata", "field_name": "Proposer",
         "value": clean_text(record.get("proposer", "Unknown")),
         "confidence": 90,
         "evidence": clean_text(record.get("proposer", ""))},
        {"category": "Content", "field_name": "Title",
         "value": record["title"], "confidence": 95,
         "evidence": record["title"]},
        {"category": "Content", "field_name": "Summary",
         "value": clean_text(record.get("summary", "")),
         "confidence": 85,
         "evidence": clean_text(record.get("summary", ""))[:500]},
        {"category": "Content", "field_name": "Products Impacted",
         "value": products_str, "confidence": products_conf,
         "evidence": str(record.get("products", ""))},
        {"category": "Assessment", "field_name": "Sector Impact",
         "value": clean_text(record.get("impact", "Unknown")),
         "confidence": 80,
         "evidence": clean_text(record.get("impact_comment", ""))},
        {"category": "Assessment", "field_name": "Likelihood",
         "value": clean_text(record.get("likelihood", "Unknown")),
         "confidence": 75,
         "evidence": clean_text(record.get("likelihood_comment", ""))},
        {"category": "Assessment", "field_name": "Status",
         "value": clean_text(record.get("status", "Unknown")),
         "confidence": 85,
         "evidence": clean_text(record.get("status", ""))},
        {"category": "Dates", "field_name": "Dates",
         "value": clean_text(record.get("dates", "")),
         "confidence": 80,
         "evidence": clean_text(record.get("dates", ""))},
        {"category": "Metadata", "field_name": "Source URL",
         "value": record.get("url", ""), "confidence": 99,
         "evidence": record.get("url", "")},
    ]


def import_record(conn, record: dict, dry_run: bool = False) -> dict:
    """Import a single Tamarind record. Returns summary dict."""
    url = record["url"]
    jurisdiction = record["jurisdiction"]
    meta = COUNTRY_META.get(jurisdiction, {"flag": "\U0001f3f3\ufe0f", "language": "EN"})
    now = datetime.now().strftime("%b %d, %Y %H:%M")

    result = {
        "url": url, "jurisdiction": jurisdiction,
        "source_created": False, "regulation_created": False,
        "fields_created": 0, "evidence_created": 0,
        "skipped": False, "updated": False,
    }

    # Check if source already exists
    existing_source = conn.execute(
        "SELECT id FROM sources WHERE external_url = ?", (url,)
    ).fetchone()

    if existing_source:
        source_id = existing_source["id"]
        result["skipped"] = True
        # Check if regulation and fields exist too
        existing_reg = conn.execute(
            "SELECT id FROM regulations WHERE source_id = ?", (source_id,)
        ).fetchone()
        if existing_reg:
            return result
    else:
        if dry_run:
            result["source_created"] = True
            result["regulation_created"] = True
            result["fields_created"] = 12
            result["evidence_created"] = 12
            return result

        # Insert source
        products_str = parse_products(record.get("products"))
        source_type_str = parse_source_type(record.get("source_type"))
        conn.execute(
            """INSERT INTO sources
               (flag, country, source_name, title, language, doc_type, discovered,
                status, stage, external_url, origin)
               VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (meta["flag"], jurisdiction, record["source_name"], record["title"],
             meta["language"], source_type_str, now,
             "Ready for Review", "Analyst Review", url, "tamarind"),
        )
        source_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        result["source_created"] = True

    # Insert regulation
    if not dry_run:
        topic = derive_topic(parse_products(record.get("products")))
        conn.execute(
            """INSERT INTO regulations
               (source_id, title, regulatory_body, jurisdiction, topic, summary,
                status, created_at, updated_at, origin)
               VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (source_id, record["title"], record["source_name"], jurisdiction,
             topic, clean_text(record.get("summary", "")),
             "Pending", now, now, "tamarind"),
        )
        reg_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        result["regulation_created"] = True

        # Insert fields and evidence
        fields = build_fields(record)
        for field_def in fields:
            conn.execute(
                """INSERT INTO regulation_fields
                   (regulation_id, source_id, category, field_name,
                    extracted_value, confidence)
                   VALUES (?,?,?,?,?,?)""",
                (reg_id, source_id, field_def["category"], field_def["field_name"],
                 field_def["value"], field_def["confidence"]),
            )
            field_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            result["fields_created"] += 1

            conn.execute(
                """INSERT INTO evidence
                   (source_id, regulation_id, field_id, excerpt, section,
                    source_reference, immutable)
                   VALUES (?,?,?,?,?,?,1)""",
                (source_id, reg_id, field_id,
                 field_def["evidence"][:1000] if field_def["evidence"] else "",
                 field_def["category"], url),
            )
            result["evidence_created"] += 1

    return result


def reset_imported(conn):
    """Delete all Tamarind-imported records, preserving seed data."""
    # Delete in dependency order
    conn.execute("""
        DELETE FROM review_decisions WHERE field_id IN (
            SELECT id FROM regulation_fields WHERE source_id IN (
                SELECT id FROM sources WHERE origin = 'tamarind'))
    """)
    conn.execute("""
        DELETE FROM evidence WHERE source_id IN (
            SELECT id FROM sources WHERE origin = 'tamarind')
    """)
    conn.execute("""
        DELETE FROM regulation_fields WHERE source_id IN (
            SELECT id FROM sources WHERE origin = 'tamarind')
    """)
    conn.execute("DELETE FROM regulations WHERE origin = 'tamarind'")
    conn.execute("DELETE FROM sources WHERE origin = 'tamarind'")
    conn.commit()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(
        description="Import Tamarind Batch 1 structured outputs into the Regulatory Intelligence database.",
    )
    parser.add_argument("--input", required=True, help="Path to Batch 1 Excel file (.xlsx)")
    parser.add_argument("--dry-run", action="store_true", help="Parse and validate without writing to database")
    parser.add_argument("--reset-imported", action="store_true",
                        help="Delete all previously imported Tamarind records before importing")
    args = parser.parse_args()

    input_path = args.input
    if not os.path.exists(input_path):
        print(f"ERROR: File not found: {input_path}")
        sys.exit(1)

    print(f"{'[DRY RUN] ' if args.dry_run else ''}Importing Tamarind data from: {input_path}")
    print()

    # Initialize database
    init_db()
    conn = get_connection()

    if args.reset_imported:
        if args.dry_run:
            count = conn.execute("SELECT COUNT(*) FROM sources WHERE origin = 'tamarind'").fetchone()[0]
            print(f"[DRY RUN] Would delete {count} previously imported sources")
        else:
            reset_imported(conn)
            print("Deleted all previously imported Tamarind records.")
        print()

    # Read Excel
    records, parse_errors = read_batch1_xlsx(input_path)
    print(f"Files discovered: 1")
    print(f"Records parsed: {len(records)}")
    print(f"Parse errors: {len(parse_errors)}")
    print()

    # Import
    summary = {
        "sources_created": 0, "regulations_created": 0,
        "fields_created": 0, "evidence_created": 0,
        "skipped": 0, "updated": 0, "errors": 0,
    }
    import_errors = []

    for record in records:
        try:
            result = import_record(conn, record, dry_run=args.dry_run)
            if result["source_created"]:
                summary["sources_created"] += 1
            if result["regulation_created"]:
                summary["regulations_created"] += 1
            summary["fields_created"] += result["fields_created"]
            summary["evidence_created"] += result["evidence_created"]
            if result["skipped"]:
                summary["skipped"] += 1
        except Exception as e:
            summary["errors"] += 1
            import_errors.append({
                "row": record.get("_row", "?"),
                "url": record.get("url", "?"),
                "error": str(e),
            })

    if not args.dry_run:
        conn.commit()
    conn.close()

    # Print summary
    print("=" * 60)
    print("IMPORT SUMMARY")
    print("=" * 60)
    print(f"  Sources imported:    {summary['sources_created']}")
    print(f"  Regulations imported:{summary['regulations_created']}")
    print(f"  Fields imported:     {summary['fields_created']}")
    print(f"  Evidence imported:   {summary['evidence_created']}")
    print(f"  Existing skipped:    {summary['skipped']}")
    print(f"  Records updated:     {summary['updated']}")
    print(f"  Validation errors:   {summary['errors']}")
    print(f"  Malformed records:   {len(parse_errors)}")
    print()

    # Write error report if any
    all_errors = parse_errors + import_errors
    if all_errors:
        error_path = Path(__file__).parent / "data" / "import_errors.json"
        error_path.parent.mkdir(parents=True, exist_ok=True)
        if not args.dry_run:
            with open(error_path, "w", encoding="utf-8") as f:
                json.dump(all_errors, f, indent=2, ensure_ascii=False)
            print(f"Error report written to: {error_path}")
        else:
            print("[DRY RUN] Errors found:")
            for e in all_errors:
                print(f"  Row {e.get('row', '?')}: {e.get('error', 'unknown')}")
    else:
        print("No errors.")

    if args.dry_run:
        print("\n[DRY RUN] No changes were made to the database.")


if __name__ == "__main__":
    main()
